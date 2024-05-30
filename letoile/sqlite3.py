import sqlite3
import openpyxl

workbook = openpyxl.load_workbook('ЛЭТУАЛЬ.xlsx')
sheet = workbook.active

with sqlite3.connect("letoile.db") as con: #Создание соединения
      cursor = con.cursor()
      cursor.execute('''
                     CREATE TABLE IF NOT EXISTS products (
                     id INTEGER PRIMARY KEY AUTOINCREMENT,
                     product TEXT,
                     current_price REAL,
                     old_price REAL,
                     company_id INTEGER,
                     rating REAL,
                     reviews_count INTEGER,
                     discount REAL,
                     FOREIGN KEY (company_id) REFERENCES companies(id))
                     ''')
      cursor.execute("""
                     CREATE TABLE IF NOT EXISTS companies (
                     id INTEGER PRIMARY KEY AUTOINCREMENT,
                     company_name TEXT)
                     """)
      
      for row in range(2, sheet.max_row + 1): #Заполнение таблицы products
             product_name = sheet.cell(row=row, column=1).value
             brand = sheet.cell(row=row, column=2).value
             current_price = sheet.cell(row=row, column=3).value
             old_price = sheet.cell(row=row, column=4).value
             discount = sheet.cell(row=row, column=5).value
             reviews_count = sheet.cell(row=row, column=6).value
             rating = sheet.cell(row=row, column=7).value

             cursor.execute("SELECT id FROM companies WHERE company_name = ?", (brand,)) #Проверка наличия компании в таблице companies
             company_id = cursor.fetchone()
             if company_id:
                     company_id = company_id[0]
             else:
                     cursor.execute("INSERT INTO companies (company_name) VALUES (?)", (brand,)) #Добавление компании в companies, если ее нет
                     company_id = cursor.lastrowid

        #Заполнение таблицы products
             cursor.execute("""
                            INSERT INTO products (product, current_price, old_price, company_id, rating, reviews_count, discount)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (product_name, current_price, old_price, company_id, rating, reviews_count, discount))
             

      print("Средняя цена продуктов по каждой фирме:\n") #Задание1
      cursor.execute("""
                     SELECT c.company_name, AVG(p.current_price) AS avg_price
                     FROM products p
                     JOIN companies c ON p.company_id = c.id
                     GROUP BY c.id
                     ORDER BY c.company_name;
                     """)
      results = cursor.fetchall()
      for row in results:
              print(f"{row[0]}, Средняя цена: {row[1]:.2f}")
              
              

      print("\nРазница между средним значением данной фирмы и ценой продукта:\n") #Задание2
      cursor.execute("""
                     SELECT DISTINCT p.product, p.current_price, (p.current_price - (SELECT AVG(current_price)
                     FROM products
                     WHERE company_id = p.company_id)) AS price_diff
                     FROM products p
                     ORDER BY p.product
                     """)
      results = cursor.fetchall()
      for row in results:
              print(f"{row[0]}, Текущая цена: {row[1]}, Разница от средней: {row[2]:.2f}")

conn.close() #Закрытие соединения